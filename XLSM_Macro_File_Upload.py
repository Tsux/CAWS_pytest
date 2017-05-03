from openpyxl import load_workbook

#sheet_ranges = wb['range names']
#print(sheet_ranges['D18'].value)

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter


import requests
import datetime
import time

target_url = open("/Users/apatel/Documents/workspace/FILE_SUBMIT_TOKEN.txt", 'a+') 

for i in range(1,2):

    wb = load_workbook(filename = '/Users/apatel/Documents/workspace/Workbook1.xlsm')
    
    #dest_filename = "/Users/apatel/Documents/workspace/empty_book_%s.xlsm" % i
    dest_filename = "empty_book_%s.xls" % i
    
    ws1 = wb.active
    ws1.title = "range names"
    
    for row in range(1, 40):
        ws1.append(range(600))
        ws2 = wb.create_sheet(title="Pi")
        ws2['F5'] = 3.14
        ws3 = wb.create_sheet(title="Data_%s" % i)
    
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
            
    print(ws3['AA10'].value)
    wb.save(filename = dest_filename)
    
    #url = "http://apibeta.nsslabs.com/Scan/file/"
    url = "https://data.nsslabs.com/Scan/file/"
    #url = "http://10.144.192.71:8081/"

    #payload = "------WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data; name=\"File\"; filename=\"" + dest_filename +  "\"\r\nContent-Type: application/vnd.openxmlformats-officedocument.wordprocessingml.document\r\n\r\n\r\n------WebKitFormBoundary7MA4YWxkTrZu0gW--"
    #payload = "------WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data;  \r\nContent-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n\r\n------WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data; name=\"template\"\r\n\r\n1\r\n------WebKitFormBoundary7MA4YWxkTrZu0gW--"
    #payload = {'Content-Disposition': "Form-data", "name" : "File", "filename" : dest_filename , "name" :"template \n\r\n\r1\n\r\n\r "} 
    #+  "\"\r\nContent-Type: application/vnd.openxmlformats-officedocument.wordprocessingml.document\r\n\r\n\r\n------WebKitFormBoundary7MA4YWxkTrZu0gW--"
    #payload = "------WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data; name=\"File\"; filename=\" %s \"\r\nContent-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n\r\n------WebKitFormBoundary7MA4YWxkTrZu0gW\r\nContent-Disposition: form-data; name=\"template\"\r\n\r\n1\r\n------WebKitFormBoundary7MA4YWxkTrZu0gW--" % dest_filename
    headers = {
        #'content-type': "multipart/form-data; boundary=----WebKitFormBoundary7MA4YWxkTrZu0gW",
        #'content-type': "Form-data",
        'authorization': "Basic cGF0ZWxsYWJzOjI5OTJBMkQ0OEVEOTQyQTJBMzU5NkQ3NjJBNDQxNUJD",
        'cache-control': "no-cache"
        ''
        }
    
    #response = requests.request("POST", url, data=payload, headers=headers)
    files = {'file': open(dest_filename, 'rb')}
    response = requests.post(url, files=files,  headers=headers, data = {"platform" : "1", "applicationPackage" : "Microsoft Office 2013"})
    now = datetime.datetime.now()
    print "Response for the http://apibeta.nsslabs.com/Scan/file/ request which returns the token : "
    print now , " ; " , dest_filename, " ; " , (response.text)
    line = response.text
    line=line.replace("Token", "")
    line=line.replace("{\"\":", "")
    line=line.replace("\"}", "")
    line=line.replace("\"", "")
    line=line.rstrip('\n')
    line=line.rstrip('\r')
    line=line.lstrip()
    print "line is :", line
    capture_token_string = str(i) + " ; " + str(now)  + " ; "  + dest_filename + " ; " + str(line)
    target_url.write(capture_token_string)
    target_url.write("\n")
    

target_url.close()
print "program is done"

