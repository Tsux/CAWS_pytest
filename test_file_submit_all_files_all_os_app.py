import pytest, requests, datetime, time, base64, json, os, pprint 
from datetime import datetime, timedelta
from CAWS_API_2_2_4 import byteify
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
#from openpyxl import Workbook
#from matplotlib.mathtext import DELTA
#sheet_ranges = wb['range names']
#print(sheet_ranges['D18'].value)

#from openpyxl import Workbook
#from openpyxl.compat import range
#from openpyxl.utils import get_column_letter



Dut = 4 
#target_url = open("/Users/apatel/Documents/workspace/URL_SUBMIT_TOKEN.txt", 'a+') 

    
if Dut == 1:
    #cawsqa -- user/password
    username = "amitpatel741234567890gmailcom"
    password = "D7A37396AB0C40CBA489D88316D727F3"
    url = "http://apiqa.qa.colo1.nsslabs.com"

if Dut == 2:
    #cawsqa -- user/password
    username = "amitpatel"
    password = "BB8FA646126D4C8991C6088E1E60E684"
    url = "http://apiqa.qa.colo1.nsslabs.com"

if Dut == 3:
    #caws qa patch
    username = "patellabs"
    password = "240DA1A652B44014993D59986156DE47"
    url = "http://apiqa-patch.qa.colo1.nsslabs.com"

if Dut == 4:
    #caws qa patch
    username = "patellabs"
    password = "2992A2D48ED942A2A3596D762A4415BC"
    url = "https://data.nsslabs.com"
    
if Dut == 5:
    #caws Beta 
    username = "patellabs"
    password = "2992A2D48ED942A2A3596D762A4415BC"
    #url = "http://10.144.192.60:82"
    #url = "https://apibeta.nsslabs.com"
    url = "https://apibeta.nsslabs.com"
    #url = "http://apibeta.nsslabs.com/Scan/url/"
    
@pytest.fixture(scope="module")
def file_submit_log():
    username = "patellabs"
    password = "2992A2D48ED942A2A3596D762A4415BC"
    url = "https://data.nsslabs.com"
    utc_datetime = datetime.utcnow()
    Today_Date_time = utc_datetime.strftime("%Y-%m-%d-%H-%M-%S")
    Today_Date = utc_datetime.strftime("%Y-%m-%d")
    
    
    #my_datetime = datetime.strptime("2017-03-14-20-57-38" , '%Y-%m-%d-%H-%M-%S')
    #delta =  (utc_datetime - my_datetime).total_seconds()
    #delta =  (utc_datetime - my_datetime).seconds
    #print "delta is:" , delta , ":"
    print "utc_datetime is:", utc_datetime, ":"
    #Today_Date_time = "2017-03-14-20-57-38"
    #Today_Date = "2017-03-14"
    #time.sleep(300)
    tmp_data_directory = "/Users/design/GitHub/CAWS_API_2.2_V1/files_for_summit_source/tmp/"+utc_datetime.strftime("%Y-%m-%d-%H-%M-%S")+"/"
    #tmp_data_directory = "/Users/apatel/Documents/workspace/CAWS_API_2.2/files_for_summit_source/tmp/2017-03-15-16-06-40/"
    if not os.path.exists(tmp_data_directory):
        os.makedirs(tmp_data_directory)
        print "new directory created . already exists"
    else:
        print "directory already exists"
        
    #Today_Date_time = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    #Today_Date = datetime.datetime.now().strftime("%Y-%m-%d")
    print "Todays date is:" , Today_Date
    print "Todays date n time  is:" , Today_Date_time
    #FILE_SUBMIT_TOKEN_FILE = open("/Users/apatel/Documents/workspace/File_Submit_Token_" + Today_Date + ".txt", 'w')
    #yield (Today_Date , Today_Date_time)
    #yield (Today_Date , Today_Date_time)
    file_list_dict = {}
    file_list_dict["1_mb_xlsx_saveas.xlsx"] = {
                                               "File_name" : "1_mb_xlsx_saveas.xlsx",
                                               "File_location" : "/Users/design/GitHub/CAWS_API_2.2_V1/files_for_summit_source/",
                                               "Submit_file_location" : tmp_data_directory,
                                               "Token_file_location" : tmp_data_directory,
                                               #"Token_file_location" : tmp_data_directory,
                                               #"Token_file_name" : tmp_data_directory+File_name,
                                               "Submit_name" : "None",
                                               "Token" : "None",
                                               "Submit_time" : "None",
                                               "Status" : "None",
                                               "NSSId" : "None",
                                               "Platform" : "Windows7",
                                               "ApplicationPackage" : "Microsoft Office 2013",
                                               "Result_time" : "None",
                                               "Md5hash" : "None",
                                               "Time_delta" : "None"
                                                 }
    start_up_config_dict = {"url": url,"username": username, "password": password, "utc_datetime": utc_datetime, "file_list_dict": file_list_dict, "tmp_data_directory": tmp_data_directory}
    return (start_up_config_dict)

#def no_test_test_01_submit_xlsx_file(file_submit_log):
def test_01_submit_xlsx_file(file_submit_log):  
    (fname,token, now, hash) = submit_file_new(file_submit_log )
    file_list_dict = file_submit_log["file_list_dict"]["1_mb_xlsx_saveas.xlsx"]
    platform = file_list_dict["Platform"]
    application = file_list_dict["ApplicationPackage"]
    assert fname != token
    assert fname != application
    assert fname != platform
    print "url is : ", url, "token is :", token, " for file name : ", fname, "submitted at :", now, " token length = ", len(token), hash, len(hash)
 
def submit_file_new(file_submit_log):
    url = file_submit_log["url"]+ '/Scan/file/'
    username = file_submit_log["username"]
    password = file_submit_log["password"]  
    Today_Date_time = str(file_submit_log["utc_datetime"].strftime("%Y-%m-%d-%H-%M-%S"))
    print "Today_Date_time is : " , Today_Date_time
    #str(utc_datetime.strftime("%Y-%m-%d-%H-%M-%S")
    file_list_dict = file_submit_log["file_list_dict"]["1_mb_xlsx_saveas.xlsx"]
    file = file_list_dict["File_name"]
    platform = file_list_dict["Platform"]
    application = file_list_dict["ApplicationPackage"]
    token = "None"
    hash = "None"
    message = "None"
    print "file name is :", file
 
    if( "xlsx" in file):
        print "file name matched."
        file = update_xls_file (file, file_submit_log)
         
 
    base64string = base64.b64encode('%s:%s' % (username, password))
    headers = {
        'authorization': "Basic %s" % base64string,
        'cache-control': "no-cache"
        }
    files = {'file': open(file, 'rb')}
    #response = requests.post(base_api_url, files=files,  headers=headers, data = {"platform" : "1", "applicationPackage" : "Adobe Reader DC 2015.017.20050"})
    response = requests.post(url, files=files,  headers=headers, data = {"platform" : platform, "applicationPackage" : application})
    #now = datetime.datetime.now()
    #now = datetime.utcnow()
    now = str(datetime.utcnow().strftime("%Y-%m-%d-%H-%M-%S"))
    file_list_dict["Submit_time"] = now
    print "Response for the ",url, "Scan/file/ request which returns the token : "
    print (str(now) + " ; " + file + " ; " + (response.text))
    line = response.text
    capture_msg_dict = json.loads(response.text)
    capture_msg_dict = byteify(capture_msg_dict)
    if "This file was already submitted" in line:
        hash = capture_msg_dict['Message'].split(" ")[-1:][0]
        message = capture_msg_dict['Message']
    elif "platform" in line:
        token = "None"
        message = capture_msg_dict['Message']
    elif "application" in line:
        token = "None"
        message = capture_msg_dict['Message']
    else:
        print "toen before is: " , file_list_dict["Token"]
        token = capture_msg_dict['Token']
        message = "File Submit Token"
        file_list_dict["Token"] = token     
        #print "this file was already submitted and its  token is:" , token
 
    if(token != "None"):
        capture_token_string = str(now) + " ; " + file + " ; " + str(token) + " ; " + message + " ; " + str(hash)
    else:
        capture_token_string = str(now) + " ; " + file + " ; " + str(token) + " ; " + message + " ; " + str(hash)
    #capture_token_string =  str(now) + " ; " + file + " ; " + str(line)
    #target_url = open("/Users/apatel/Documents/workspace/File_Submit_Token.txt", 'a+') 
    FILE_SUBMIT_TOKEN_FILE = open("/Users/apatel/Documents/workspace/File_Submit_Token_"+Today_Date_time+".txt", 'a+')
    FILE_SUBMIT_TOKEN_FILE.write(capture_token_string)
    FILE_SUBMIT_TOKEN_FILE.write("\n")
    print "dictionary for file is:",   file_list_dict
    json.dump(file_list_dict, open(file_submit_log['tmp_data_directory']+file_list_dict["File_name"]+'.txt','w'))
     
    FILE_SUBMIT_TOKEN_FILE.close()
#    token_file.write(capture_token_string)
    return (file, token, now, hash)
 
 
#@pytest.mark.parametrize("test_input, expected_output",
@pytest.mark.parametrize("file_name, platform, application",
                         [
                            # Microsoft Office with windows 7. 
#                             ("DOC_doc.doc",'Windows7', 'Microsoft Office 2013'),
#                             ("DOC_docx.docx",'Windows7', 'Microsoft Office 2013'),
#                            ("1_mb_xls_save_as.xls",'Windows7', 'Microsoft Office 2013'),
                             ("1_mb_xlsx_saveas.xlsx",'Windows7', 'Microsoft Office 2013'),                            
#                             ("DOC_csv.csv",'Windows7', 'Microsoft Office 2013'),
#                             ("DOC_pps.pps.ppt", 'Windows7', 'Microsoft Office 2013'),
#                             ("1_mb_ppt_saveass.pptx", 'Windows7', 'Microsoft Office 2013'),
#                             #pdf file
#                             ("DOC_pdf.pdf", "Windows7", "Adobe Reader DC 2015.017.20050"),
#                             ("DOC_pdf.pdf","Windows7" , "Adobe Reader DC 2015.007.20033"),
#                             ("DOC_pdf.pdf","Windows7" , "Adobe Reader 9.4"),
#                             ("DOC_pdf.pdf","Windows7" , "Adobe Reader DC 2015.020.20039"),                            
#                             #Apple files with Win 7
#                             ("SampleAudio_0.5mb.mp3",'Windows7', 'Quicktime 7.79' ),
#                             ("SampleAudio_0.5mb.mp3",'Windows7', 'Itunes 12.5.1' ),
#                             ("SampleAudio_0.5mb.mp3",'Windows7', 'Itunes 12.5.4' ),                            
#                             #VLC
#                             ("SampleAudio_0.5mb.mp3",'Windows7', 'VLC 2.2.3' ),
#                             ("SampleAudio_0.5mb.mp3",'Windows7', 'VLC 2.2.4' )
                        ]
                         )
 
#sample token = 5008F26FECE31FFCA09D44CEF2CE28B9
#def test_submit_pdf(test_input, expected_output):
# def test_submit_file(file_name, platform, application,file_submit_log):
#     #FILE_SUBMIT_TOKEN_FILE = file_submit_log[2]
#     #(fname,token, now) = submit_pdf(url, username, password, "/Users/apatel/Documents/CAWS/CAWS 2.2/CAWS_2.2_Release_Notes.pdf")
#     (fname,token, now, hash) = submit_file(url, username, password, file_name, platform, application,file_submit_log )
#     assert file_name != token
#     assert file_name != application
#     #assert (file_name + ";" + token) != (platform + ";" + application)
#       
#     print "url is : ", url, "token is :" , token ,"for file name : ", file_name , "submitted at :" , now , "token length = " , len(token), hash , len(hash)
#     #return (token, filename)
 
 
 
     
#def submit_file(server_api_url, username, password,file, platform, application, ):
def submit_file(server_api_url, username, password, file, platform, application, file_submit_log):
    Today_Date_time = file_submit_log[1]
    file_list_dict = file_submit_log["file_list_dict"]
    file = file
    platform = platform
    application = application
    token = "None"
    hash = "None"
    message = "None"
 
    base64string = base64.b64encode('%s:%s' % (username, password))
    base_api_url = server_api_url + '/Scan/file/'
    #print "file is :" , file
    #wb = load_workbook(filename = '/Users/apatel/Documents/workspace/Workbook1.xlsm')
    headers = {
        'authorization': "Basic cGF0ZWxsYWJzOjI5OTJBMkQ0OEVEOTQyQTJBMzU5NkQ3NjJBNDQxNUJD",
        'cache-control': "no-cache"
        }
    files = {'file': open(file, 'rb')}
    #response = requests.post(base_api_url, files=files,  headers=headers, data = {"platform" : "1", "applicationPackage" : "Adobe Reader DC 2015.017.20050"})
    response = requests.post(base_api_url, files=files,  headers=headers, data = {"platform" : platform, "applicationPackage" : application})
    now = datetime.datetime.now()
    print "Response for the ",url, "Scan/file/ request which returns the token : "
    print (str(now) + " ; " + file + " ; " + (response.text))
    line = response.text
    capture_msg_dict = json.loads(response.text)
    capture_msg_dict = byteify(capture_msg_dict)
    if "This file was already submitted" in line:
        hash = capture_msg_dict['Message'].split(" ")[-1:][0]
        message = capture_msg_dict['Message']
    elif "platform" in line:
        token = "None"
        message = capture_msg_dict['Message']
    elif "application" in line:
        token = "None"
        message = capture_msg_dict['Message']
    else:
        token = capture_msg_dict['Token']
        message = "File Submit Token"
        #print "this file was already submitted and its  token is:" , token
 
    if(token != "None"):
        capture_token_string =  str(now)  + " ; "  + file + " ; " + str(token) + " ; " + message + " ; " + str(hash)
    else:
        capture_token_string =  str(now)  + " ; "  + file + " ; " + str(token) + " ; " + message + " ; " + str(hash)
    #capture_token_string =  str(now)  + " ; "  + file + " ; " + str(line)
    #target_url = open("/Users/apatel/Documents/workspace/File_Submit_Token.txt", 'a+') 
    FILE_SUBMIT_TOKEN_FILE = open("/Users/design/GitHub/CAWS_API_2.2_V1/File_Submit_Token_"+Today_Date_time+".txt", 'a+')
    FILE_SUBMIT_TOKEN_FILE.write(capture_token_string)
    FILE_SUBMIT_TOKEN_FILE.write("\n")
     
    FILE_SUBMIT_TOKEN_FILE.close()
#    token_file.write(capture_token_string)
    return (file, token, now, hash)
 
 
'''
def update_word_doc_file (filename, new_file_name_with_location):
    #from docx import *
    document = opendocx("document.doc")
    body = document.xpath('/w:document/w:body', namespaces=nsprefixes)[0]
    body.append(paragraph('Appending this.'))
    savedocx()
'''
     
def update_xls_file (filename, file_submit_log):
    print "inside xls update"
    #Today_Date_time = str(Today_Date_time)
    Today_Date_time = str(file_submit_log["utc_datetime"].strftime("%Y-%m-%d-%H-%M-%S"))
    #wb = load_workbook(filename = "/Users/apatel/Documents/workspace/CAWS_API_2.2/files_for_summit_source/"+filename)
    wb = load_workbook(filename = "/Users/design/GitHub/CAWS_API_2.2_V1/files_for_summit_source/"+filename)
    #dest_filename = file_submit_log['tmp_data_directory']+filename[:-5]+"-"+Today_Date_time+".xlsx"
    dest_filename = file_submit_log['tmp_data_directory']+filename
    ws1 = wb.active
    ws1.title = "range names"    
    for row in range(1, 40):
        ws1.append(range(600))
        ws2 = wb.create_sheet(title="Pi")
        ws2['F5'] = 3.14
        ws3 = wb.create_sheet(title="Data_%s" %  Today_Date_time)
     
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
             
    #print(ws3['AA10'].value)
    #wb.save(filename = dest_filename)
    wb.save(filename=dest_filename)
    print "returning of xls update."
    return (dest_filename)
 
# AA
# Response for the http://apibeta.nsslabs.com/Scan/file/ request which returns the token : 
# 2017-02-16 11:39:53.240196  ;  empty_book_1.xls  ;  {"Token":"DBBAE6783EBA46428C9C9AD2BB165BE7"}
# line is : DBBAE6783EBA46428C9C9AD2BB165BE7
# program is done
 
 
'''
def update_pdf_doc_file (filename, new_file_name_with_location):
 
    from pyPdf import PdfFileWriter, PdfFileReader
    import StringIO
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
     
    packet = StringIO.StringIO()
    # create a new PDF with Reportlab
    can = canvas.Canvas(packet, pagesize=letter)
    can.drawString(10, 100, "Hello world")
    can.save()
     
    #move to the beginning of the StringIO buffer
    packet.seek(0)
    new_pdf = PdfFileReader(packet)
    # read your existing PDF
    existing_pdf = PdfFileReader(file("original.pdf", "rb"))
    output = PdfFileWriter()
    # add the "watermark" (which is the new pdf) on the existing page
    page = existing_pdf.getPage(0)
    page.mergePage(new_pdf.getPage(0))
    output.addPage(page)
    # finally, write "output" to a real file
    outputStream = file("destination.pdf", "wb")
    output.write(outputStream)
    outputStream.close()
'''
 
def test_02_check_xlsx_file_submit_status(file_submit_log):
    (Status, NSSId, delta, token) = check_xlsx_file_submit_status(file_submit_log)
    __tracebackhide__ = True
    if (str(Status) != "2"):
        pytest.fail("Status is still not completed even after wait period is %d secs and status is still : %s" %(delta, Status,))
    file_list_dict = file_submit_log["file_list_dict"]["1_mb_xlsx_saveas.xlsx"]
    platform = file_list_dict["Platform"]
    application = file_list_dict["ApplicationPackage"]
    assert token != ""
    assert application != ""
    assert platform != ""
    assert Status != ""
    assert NSSId != ""
    assert delta != ""
    assert token != ""
     
    #print "url is : ", url, "token is :" , token ,"for file name : ", fname , "submitted at :" , now , "token length = " , len(token), hash , len(hash)
     
 
def check_xlsx_file_submit_status(file_submit_log):
    pprint.pprint(file_submit_log)
    time.sleep(500)
    username = file_submit_log["username"]
    password = file_submit_log["password"]  
    #Today_Date_time = str(file_submit_log["utc_datetime"].strftime("%Y-%m-%d-%H-%M-%S")) 
    #Today_Date_time = "2017-03-14-20-57-38"   
    file_list_dict = file_submit_log["file_list_dict"]["1_mb_xlsx_saveas.xlsx"]
    url_token_list_file = "/Users/design/GitHub/CAWS_API_2.2_V1/URL_SUBMIT_TOKEN.txt"
    file_submit_result_dict = json.load(open(file_submit_log['tmp_data_directory']+file_list_dict["File_name"]+'.txt'))
    print file_submit_result_dict
    token = file_submit_result_dict["Token"]
    Submit_time = file_submit_result_dict["Submit_time"]
    file_list_dict["Submit_time"] = file_submit_result_dict["Submit_time"]
    print "submit time is :", Submit_time
    url = file_submit_log["url"]+ '/Scan/Status/file/'+str(token)
    print "url and token for the query is :", url, ";", token
    #url = "http://apibeta.nsslabs.com/Scan/Status/url/"+str(token)
    base64string = base64.b64encode('%s:%s' % (username, password))
    headers = {
        'authorization': "Basic %s" % base64string,
        'cache-control': "no-cache"
        }
    my_datetime = datetime.strptime(Submit_time , '%Y-%m-%d-%H-%M-%S')
    delta = (datetime.utcnow() - my_datetime).seconds
#     
#     if (delta > 180):
#         print "submit time to check time diff is :" , delta
#         response = requests.request("GET", url, headers=headers)
#         
#     else:
#         time.sleep(180 - delta)
#         response = requests.request("GET", url, headers=headers)
#         print "submit time to check time diff is :" , delta
#     
#     print "response is :" , response    
#     
    response_dict = {}
#   response_dict = json.loads(response.text)
#   response_dict = byteify(response_dict)
     
    #if("NSSId" not in response_dict):
    while("NSSId" not in response_dict):
       del response_dict 
       print "even after " , delta, "seconds, status is not 2. Will wait for 5 mins now"
       delta =  (datetime.utcnow() - my_datetime).seconds
       #time.sleep(300 - delta)
       time.sleep(5)
        
       response = requests.request("GET", url, headers=headers)
       #print "submit time to check time diff is :" , delta 
       response_dict = {}
       response_dict = json.loads(response.text)
       response_dict = byteify(response_dict)
         
    print "response to status check is :" , response.text
    if("NSSId" in response_dict):
        file_list_dict["NSSId"] = response_dict["NSSId"]
        file_list_dict["Status"] = response_dict['Status']
    else:
        file_list_dict["Status"] = response_dict['Status']
 
    json.dump(file_list_dict, open(file_submit_log['tmp_data_directory']+file_list_dict["File_name"]+'.txt','w'))
    print "file_list_dict is :"
    pprint.pprint(file_list_dict)
    #print url_to_scan , " ; " , token , " ; " , NSSId , " ; " , Status
    capture_token_string = url + " ; " + token + " ; " + str(file_list_dict["NSSId"])
    target_url = open("/Users/design/GitHub/CAWS_API_2.2_V1/URL_TOKEN_with_NSSId.txt", 'a+')
    target_url.write(str(capture_token_string))
    target_url.write("\n")
    target_url.close()
    #print "Response for the url scan status http://apibeta.nsslabs.com/Scan/Status/url/{token=634DF6E7AC6742B187EC945871EBE4DC} request which returns NSS ID and status."
    return(file_list_dict["Status"], file_list_dict["NSSId"], (datetime.utcnow() - my_datetime).seconds, token )
 
print "program is done."

def test_03_check_xlsx_file_NSSId_details(file_submit_log):
    NSSId_dict = {} 
    NSSId_dict = check_xlsx_file_NSSId_details(file_submit_log)
    #__tracebackhide__ = True
    for key in NSSId_dict:
        assert NSSId_dict[key] != ""
        print "key is :" , key , " and values is:", NSSId_dict[key] #ErrorMessage
        if (key == "ErrorMessage"):
            pytest.fail("Got Errormessage %s that %s at : %s" %(key, NSSId_dict[key], NSSId_dict["ErrorDate"]))
    file_list_dict = file_submit_log["file_list_dict"]["1_mb_xlsx_saveas.xlsx"]
    platform = file_list_dict["Platform"]
    application = file_list_dict["ApplicationPackage"]
    
def check_xlsx_file_NSSId_details(file_submit_log):

    
    username = file_submit_log["username"]
    password = file_submit_log["password"]  
    #Today_Date_time = str(file_submit_log["utc_datetime"].strftime("%Y-%m-%d-%H-%M-%S")) 
    #Today_Date_time = "2017-03-14-20-57-38"   
    file_list_dict = file_submit_log["file_list_dict"]["1_mb_xlsx_saveas.xlsx"]
    url_token_list_file = "/Users/apatel/Documents/workspace/URL_SUBMIT_TOKEN.txt"
    file_submit_result_dict = json.load(open(file_submit_log['tmp_data_directory']+file_list_dict["File_name"]+'.txt'))
    Submit_time = file_submit_result_dict["Submit_time"]
    NSSId = file_submit_result_dict["NSSId"]
    Submit_time = file_submit_result_dict["Submit_time"]
    url = file_submit_log["url"]+ '/Users/Files/'+str(NSSId)
    print "url and token for the query is :" , url , ";" , NSSId
    #url = "http://apibeta.nsslabs.com/Scan/Status/url/"+str(token)
    base64string = base64.b64encode('%s:%s' % (username, password))
    headers = {
        'authorization': "Basic %s" % base64string,
        'cache-control': "no-cache"
        }
    response_dict = {}
    my_datetime = datetime.strptime(Submit_time , '%Y-%m-%d-%H-%M-%S')
    delta =  (datetime.utcnow() - my_datetime).seconds
    
    while("captureModel" not in response_dict):
        #del response_dict 
        print "even after " , delta, "seconds, NSSId details are not available. Will wait for 5 secs now"
        delta =  (datetime.utcnow() - my_datetime).seconds
        #time.sleep(300 - delta)
        time.sleep(5)
        response = requests.request("GET", url, headers=headers)    
        print "response code is:" , response, "and text is:" , response.text
        response_dict = {}
        response_dict = json.loads(response.text)
        response_dict = byteify(response_dict)

    response_dict = byteify(response_dict["captureModel"])
    pprint.pprint(response_dict)
    file_list_dict["IsMalicious"] = response_dict["IsMalicious"]
    file_list_dict["DetectionDate"] = response_dict["DetectionDate"]
    file_list_dict["Result_time"] = datetime.utcnow().strftime("%Y-%m-%d-%H-%M-%S")
    file_list_dict['Time_delta'] = delta
    file_list_dict["NSSId"] = file_submit_result_dict["NSSId"]
    json.dump(file_list_dict, open(file_submit_log['tmp_data_directory']+file_list_dict["File_name"]+'.txt','w'))
    #print "this is what will be returned :"
    #pprint.pprint (response_dict)
    return response_dict



#     file_submit_log = {'file_list_dict': {'1_mb_xlsx_saveas.xlsx': {'ApplicationPackage': 'Microsoft Office 2013',
#                                               'File_location': '/Users/apatel/Documents/workspace/CAWS_API_2.2/files_for_summit_source/',
#                                               'File_name': '1_mb_xlsx_saveas.xlsx',
#                                               'Md5hash': 'None',
#                                               'NSSId': 'None',
#                                               'Platform': 'Windows7',
#                                               'Result_time': 'None',
#                                               'Status': 'None',
#                                               'Submit_file_location': '/Users/apatel/Documents/workspace/CAWS_API_2.2/files_for_summit_source/tmp/2017-03-21-22-58-20/',
#                                               'Submit_name': 'None',
#                                               'Submit_time': '2017-03-21-22-58-29',
#                                               'Time_delta': 'None',
#                                               'Token': 'CAD9C60544BA443193238AAD635A5C33',
#                                               'Token_file_location': '/Users/apatel/Documents/workspace/CAWS_API_2.2/files_for_summit_source/tmp/2017-03-21-22-58-20/'}},
#  'password': '2992A2D48ED942A2A3596D762A4415BC',
#  'tmp_data_directory': '/Users/apatel/Documents/workspace/CAWS_API_2.2/files_for_summit_source/tmp/2017-03-21-21-18-41/',
#  'url': 'https://data.nsslabs.com',
#  'username': 'patellabs',
#  'utc_datetime': 'datetime.datetime(2017, 3, 21, 22, 44, 17, 416704)'}
    

